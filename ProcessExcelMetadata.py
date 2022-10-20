import openpyxl as xl
from openpyxl.cell.read_only import EmptyCell



iden_file_names  = ['11_IDEN2016.TXT', '11_IDEN2017.TXT', '11_IDEN2018.TXT', '11_IDEN2019.TXT']
renta_file_names = ['2_Renta2016.txt', '2_Renta2017.txt', '2_Renta2018.txt', '2_Renta2019.txt']
renta_imputacion_file_names = ['3_RentaImputacion2016.txt', '3_RentaImputacion2017.txt', '3_RentaImputacion2018.txt', '3_RentaImputacion2019.txt']
irpf_file_names = ['4_IRPF2016.txt', '4_IRPF2017.txt', '4_IRPF2018.txt', '4_IRPF2019.txt']
patrimonio_file_names = ['5_Patrimonio2016.txt', '5_Patrimonio2017.txt', '5_Patrimonio2018.txt', '5_Patrimonio2019.txt']
mod714_file_names = ['6_M714_2016.txt', '6_M714_2017.txt', '6_M714_2018.txt', '6_M714_2019.txt']
mod790_file_names = ['7_M190_2016.txt', '7_M190_2017.txt', '7_M190__2019.txt', '7_M190_2019.txt']
irpf_rrii_file_names = ['4_IRPF2016_RRII.txt', '4_IRPF2017_RRII.txt', '4_IRPF2018_RRII.txt', '4_IRPF2019_RRII.txt']
irpf_aaee_ed_file_names = ['9_IRPF2016_AAEE_ED.txt', '9_IRPF2017_AAEE_ED.txt', '9_IRPF2018_AAEE_ED.txt', '9_IRPF2019_AAEE_ED.txt']
irpf_aaee_eobj_file_names = ['9_IRPF2016_AAEE_EOBJ.txt', '9_IRPF2017_AAEE_EOBJ.txt', '9_IRPF2018_AAEE_EOBJ.txt', '9_IRPF2019_AAEE_EOBJ.txt']
irpf_aaee_eobja_file_names = ['9_IRPF2016_AAEE_EOBJA.txt', '9_IRPF2017_AAEE_EOBJA.txt', '9_IRPF2018_AAEE_EOBJA.txt', '9_IRPF2019_AAEE_EOBJA.txt']


inputExcel = './doc/00_DiseñoRegistro.xlsx'
output_folder = './out/'




def getStartColumn(worksheet, cell_content):
    for current_row in worksheet.iter_rows(min_row=0, max_row=worksheet.max_row):
        # check EmptyRow and EmptyCell
        if current_row is not None and len(current_row) > 0:
            num_col = 0
            for cell in current_row:
                if cell is not None and type(cell) is not EmptyCell:
                    if cell_content in str(cell.value):
                        return num_col
                num_col += 1
    return -1

def getStartRow(worksheet, start_col):
    num_row = 0
    for current_row in worksheet.iter_rows(min_row=0, max_row=worksheet.max_row):
        # check EmptyRow and EmptyCell
        if current_row is not None and len(current_row) > 0:
            num_col = 0
            for cell in current_row:
                if num_col >= start_col:
                    if cell is not None and type(cell) is not EmptyCell:
                        if 'Posición inicial' in str(cell.value):
                            return num_row
                num_col += 1
        num_row += 1
    return -1


def processMetadataRow(current_row, start_col):
    #{'var_name': 'IDENPER', 'var_pos_inicial': 1, 'var_tipo': 'Num', 'var_long': 11, 'var_decimales': None, 'var_desc': 'Identificador único de persona'}

    metadataMap = {}

    var_name = current_row[start_col+1]
    metadataMap['var_name'] = var_name.value

    var_pos_inicial = current_row[start_col]
    metadataMap['var_pos_inicial'] = var_pos_inicial.value

    var_tipo = current_row[start_col + 2]
    metadataMap['var_tipo'] = var_tipo.value

    var_long = current_row[start_col + 3]
    metadataMap['var_long'] = var_long.value

    var_decimales = current_row[start_col + 4]
    metadataMap['var_decimales'] = var_decimales.value

    var_desc = current_row[start_col + 5]
    metadataMap['var_desc'] = var_desc.value

    if var_tipo.value is None or var_name.value is None or var_tipo.value is None or var_long.value is None:
        return None

    return metadataMap


def processMetadata(worksheet, start_col, start_row):
    metadata = []
    num_row = 0
    for current_row in worksheet.iter_rows(min_row=0, max_row=worksheet.max_row):
        cell = current_row[start_col]
        if cell is not None and type(cell) is not EmptyCell and len(str(cell.value)) > 0 and num_row > start_row:
            metadataRow = processMetadataRow(current_row, start_col)
            if metadataRow is not None:
                metadata.append(metadataRow)
        num_row += 1
    return metadata


def getType(type_name):
    if type_name == 'Num':
        return 'NUMERIC'
    elif type_name == 'Char':
        return 'VARCHAR'
    return 'VARCHAR'

def getCanonicalName(file_name):
    offset = file_name.find('_') + 1
    file_name = file_name[offset:]
    file_name = file_name[:-4]
    return file_name


def writeLoadData(metadata, original_file_name):
    file_name = getCanonicalName(original_file_name)
    file_name_def = output_folder + 'load_' + file_name + '.sql'
    tablename = 'tbl_' + file_name

    with open(file_name_def, 'w') as f:
        f.write('USE test;\n\n')
        f.write('LOAD DATA LOCAL INFILE \'' + original_file_name + '\'\n')
        f.write('INTO TABLE ' + tablename + '\n')
        f.write('(@row)\n')
        f.write('SET \n')

        for metadata_item in metadata:
            var_name = metadata_item.get('var_name')
            var_pos_inicial = metadata_item.get('var_pos_inicial')
            var_long = metadata_item.get('var_long')

            if metadata_item == metadata[-1]:
                f.write('\t' + var_name + ' = TRIM(SUBSTR(@row,' + str(var_pos_inicial) + ', ' + str(var_long) + '))\n')
            else:
                f.write('\t' + var_name + ' = TRIM(SUBSTR(@row,' + str(var_pos_inicial) + ', ' + str(var_long) + ')),\n')

        f.write(';')



def writeCreateTable(metadata, original_file_name):
    file_name = getCanonicalName(original_file_name)
    file_name_def = output_folder + 'create_table_' + file_name + '.sql'
    tablename = 'tbl_' + file_name

    with open(file_name_def, 'w') as f:
        f.write('USE test;\n\n')
        f.write('DROP TABLE IF EXISTS ' + tablename + ';\n\n')
        f.write('CREATE TABLE ' + tablename + '(\n')

        for metadata_item in metadata:
            var_name = metadata_item.get('var_name')
            var_tipo = getType(metadata_item.get('var_tipo'))
            var_long = metadata_item.get('var_long')
            var_decimales = metadata_item.get('var_decimales')

            strLength = ''
            if var_decimales is not None:
                strLength = str(var_long) + ','+ str(var_decimales)
            else:
                strLength = str(var_long)

            if metadata_item == metadata[-1]:
                f.write('\t' + var_name + ' ' + var_tipo + '(' + strLength + ') DEFAULT NULL\n')
                # TODO var_desc
            else:
                f.write('\t' + var_name + ' ' + var_tipo + '(' + strLength + ') DEFAULT NULL,\n')

        f.write(') engine=columnstore CHARSET=latin1 COLLATE=latin1_spanish_ci;')


def getStartColRow(worksheet, file_name):
    start_col = getStartColumn(worksheet, file_name)
    if start_col < 0:
        print('ERROR reading header (start_col): ' + file_name)
    start_row = getStartRow(worksheet, start_col)
    if start_row < 0:
        print('ERROR reading header (start_col): ' + file_name)

    # print(' start col:' + str(start_col))
    # print(' start row:' + str(start_row))

    return start_col, start_row


def buildScriptFiles(file_name):
    start_col, start_row = getStartColRow(worksheet, file_name)
    if start_col < 0 or start_row < 0:
        return -1
    metadata = processMetadata(worksheet, start_col, start_row)
    writeCreateTable(metadata, file_name)
    writeLoadData(metadata, file_name)
    return 0

# process worksheets
workbook = xl.load_workbook(inputExcel, keep_links=False, read_only=True, data_only=True)

for worksheet in workbook.worksheets:
    print('Sheet: ' + worksheet.title)
    if worksheet.title == '1_IDEN':
        for iden_file_name in iden_file_names:
            if buildScriptFiles(iden_file_name) < 0:
                continue

    elif worksheet.title == '2_Renta':
        for renta_file_name in renta_file_names:
            if buildScriptFiles(renta_file_name) < 0:
                continue

    elif worksheet.title == '3_RentaImputación':
        for renta_imputacion_file_name in renta_imputacion_file_names:
            if buildScriptFiles(renta_imputacion_file_name) < 0:
                continue

    elif worksheet.title == '4_IRPF':
        for irpf_file_name in irpf_file_names:
            if buildScriptFiles(irpf_file_name) < 0:
                continue

    elif worksheet.title == '5_Patrimonio':
        for patrimonio_file_name in patrimonio_file_names:
            if buildScriptFiles(patrimonio_file_name) < 0:
                continue

    elif worksheet.title == '6_Mod714':
        for mod714_file_name in mod714_file_names:
            if buildScriptFiles(mod714_file_name) < 0:
                continue

    elif worksheet.title == '7_Mod190':
        for mod790_file_name in mod790_file_names:
            if buildScriptFiles(mod790_file_name) < 0:
                continue

    elif worksheet.title == '8_IRPF_RRII':
        for irpf_rrii_file_name in irpf_rrii_file_names:
            if buildScriptFiles(irpf_rrii_file_name) < 0:
                continue

    elif worksheet.title == '9_IRPF_AAEE_ED':
        for irpf_aaee_ed_file_name in irpf_aaee_ed_file_names:
            if buildScriptFiles(irpf_aaee_ed_file_name) < 0:
                continue

    elif worksheet.title == '9_IRPF_AAEE_EOBJ':
        for irpf_aaee_eobj_file_name in irpf_aaee_eobj_file_names:
            if buildScriptFiles(irpf_aaee_eobj_file_name) < 0:
                continue

    elif worksheet.title == '9_IRPF_AAEE_EOBJA':
        for irpf_aaee_eobja_file_name in irpf_aaee_eobja_file_names:
            if buildScriptFiles(irpf_aaee_eobja_file_name) < 0:
                continue












